# This program creates Tetbury forms and includes GP and Insurance details
# V1 Runs on Alisons PC
#
# V1.1 01/04/2022
# Update to reflect the changes made by WriteUpp in the export activities and file (NHS number

#v1.2 16/05/2022
# Change in time field to refelct WritrUpp change

#v1.2a 24/5/22
#Only changes to make it work on Alison's PC

import os
import openpyxl
import csv
import time
import glob
from selenium import webdriver
from selenium.webdriver.support.select import Select
from tkinter import *
import pymsgbox


def loginWriteupp():
    #testfield = entrySuffix.get()
    #Login to writeUpp
    loginDriver = driver
    loginDriver.get(loginURL)
    time.sleep(2)
    userNameField = loginDriver.find_element_by_id('EmailAddress')
    userNameField.send_keys(userName)
    passwordField = driver.find_element_by_id('Password')
    passwordField.send_keys(password)
    time.sleep(1)
    submitButton = driver.find_element_by_xpath('/html/body/div[2]/main/div/div[2]/div/form/div[3]/div/div/button')
    submitButton.click()
    time.sleep(5)

def getInsuranceCompanies():
    #Find all the third party insurance companies and export them and then read into a list
    driver.get(thirdURL)
    time.sleep(1)
    insurerSelect = driver.find_element_by_xpath('/html/body/form/div[5]/div/div[3]/div[2]/div/div[1]/select')
    Select(insurerSelect).select_by_visible_text('Insurer')
    time.sleep(2)
    exportCSV = driver.find_element_by_xpath('/html/body/form/div[5]/div/div[4]/div/div/div/div/div/div[3]/a')
    exportCSV.click()
    time.sleep(2)
    os.chdir(wd)
    global companies
    companies = [] #this list contains all the insurance companies in use.
    thirdPartiesFile = open('ThirdParties.csv')
    csv_insurers = csv.reader(thirdPartiesFile)
    for row in csv_insurers:
        companies.append(row[0])
    companies.pop(0) # get rid of header row
    thirdPartiesFile.close()

def process_patients():
    os.chdir(wd)
    with open(activity_filename) as p:
        patients = csv.reader(p)
        next(patients) # skip header row
        for patient in patients:
            tp_name = patient[1]
            tp_appointment_time = patient[4]
            tp_appointment_date = patient[6]
            tp_appointment_type = patient[2]
            wuid = patient[0]
            searchField = driver.find_element_by_id('ctl00_ctl00_Content_siteHead_dfSearchWidget')
            searchField.send_keys(wuid)
            driver.find_element_by_id('ctl00_ctl00_Content_siteHead_btnSearch').click()
            time.sleep(1)
            age_dob_field = driver.find_element_by_id('ctl00_ctl00_Content_ContentPlaceHolderPS_dateOfBirth')
            age_dob = age_dob_field.text
            tp_DOB = age_dob[:9]
            tp_age = age_dob[age_dob.find("(")+1:age_dob.find(")")] #grab text between the brackets
            tp_address = driver.find_element_by_id('ctl00_ctl00_Content_ContentPlaceHolderPS_dfFullAddress').text
            tp_home_phone = driver.find_element_by_xpath('/html/body/form/div[5]/div[3]/div/div/div[2]/article[1]/table/tbody/tr[1]/td/span/a').text
            tp_mobile = driver.find_element_by_xpath('/html/body/form/div[5]/div[3]/div/div/div[2]/article[1]/table/tbody/tr[3]/td/span/a').text
            #tp_email = driver.find_element_by_xpath('/html/body/form/div[5]/div[3]/div/div/div[2]/article[1]/table/tbody/tr[4]/td/span/a').text
            tp_nhs = driver.find_element_by_xpath('/html/body/form/div[5]/div[3]/div/div/div[1]/article[1]/table/tbody/tr[8]/td/div/p').text

            #Get GP and insurance details
            third_parties = driver.find_elements_by_class_name('patient-summary__third-parties__name')
            thirdparty_attributes = driver.find_elements_by_class_name('patient-summary__third-parties__attribute')
            #set theses tom blank in case they don't exist
            tp_insurance_co = ''
            tp_insurance_co_address =''
            tp_membership_no =''


            tp_gp_surgery =''
            tp_authorisation =''
            for third_party in third_parties:
                third_party_text = third_party.text
                first_word = third_party_text.split(' ', 1)[0]
                back_string = third_party_text.split("- ", 1)[1]
                doctor_word = back_string.split(' ', 1)[0]
                if first_word in companies:
                    # print('Found an insurance company')
                    # print(third_party_text)
                    tp_insurance_co = third_party_text.split('-', 1)[0]
                    tp_insurance_co_address = back_string
                elif doctor_word in dr_list:
                    # print('found a Doctor')
                    # print(third_party_text)
                    tp_gp_name = back_string.split(',', 1)[0]
                    tp_gp_surgery = third_party_text.replace(tp_gp_name+",","") #remove doctors name
                else:
                     print('You have found something else')
            #Check for policy number and autorisation code
            if thirdparty_attributes != []:
                # print('something here')
                for attribute in thirdparty_attributes:
                    third_party_attribute_text = attribute.text
                    if "Policy Number" in third_party_attribute_text:
                        tp_membership_no = third_party_attribute_text.split(':', 1)[1] #get the text after the :
                    if "Authorisation Code" in third_party_attribute_text:
                        tp_authorisation = third_party_attribute_text.split(':', 1)[1] #get the text after the :

            # Start creating the spreadsheet for this patient
            os.chdir(wd)
            clinic_file = openpyxl.load_workbook(template_file)
            tp_ws = clinic_file.active

            # populate the spreadsheet
            tp_ws.cell(16, 3).value = tp_name.split(' ', 1)[0]
            tp_ws.cell(17, 3).value = tp_name.split(' ', 1)[1]
            tp_ws.cell(20, 3).value = tp_age
            tp_ws.cell(19, 3).value = tp_DOB
            # tp_ws.cell(20, 7).value = tp_email
            tp_ws.cell(18, 3).value = tp_nhs
            tp_ws.cell(21, 3).value = tp_address
            tp_ws.cell(22, 3).value = tp_home_phone
            tp_ws.cell(23, 3).value = tp_mobile
            tp_ws.cell(25, 2).value = tp_gp_name
            tp_ws.cell(26, 2).value = tp_gp_surgery
            tp_ws.cell(31, 4).value = tp_appointment_type
            tp_ws.cell(35, 4).value = tp_appointment_date + ' ' + tp_appointment_time
            tp_ws.cell(65, 4).value = tp_insurance_co
            tp_ws.cell(71, 4).value = tp_insurance_co_address
            tp_ws.cell(67, 4).value = tp_membership_no
            tp_ws.cell(69, 4).value = tp_authorisation
            tp_ws.cell(77, 3).value = tp_appointment_date
            tp_ws.cell(79, 3).value = tp_appointment_time

            # save and close
            os.chdir(this_dir)
            this_filename = tp_name+'.xlsx'
            clinic_file.save(this_filename)
            clinic_file.close




def finishUp():
    root.destroy()
    # Delete all oldfiles
    oldFiles = glob.glob(wd + '//*.csv')
    for f in oldFiles:
        os.remove(f)
    return

def alldone():
    pymsgbox.alert('All done')
    driver.quit()
    return

def getActivity():
    #Export all the Activity for the given dates in WriteUpp

    #Delete all oldfiles
    oldFiles = glob.glob(wd+'//*.csv')
    for f in oldFiles:
        os.remove(f)

    # Get the Acivity report
    root.withdraw()
    driver.get(activityURL)
    time.sleep(1)
    pymsgbox.alert('Enter Dates and  click OK')

    time.sleep(2)
    activity_BTN = driver.find_element_by_id('ctl00_ctl00_Content_ContentPlaceHolder1_btnExportCsv')
    activity_BTN.click()
    os.chdir(wd)
    os.rename(wu_activity_filename,activity_filename)

def setup_folder():
    folder_name = entryFolder.get()
    global this_dir
    this_dir = os.path.join(HospitalSheetDirectory, folder_name)
    if not os.path.exists(this_dir):
        os.mkdir(this_dir)
    return

version_no = "1.2a Alison 24/05/2022"
writeUppURL = 'https://dr-emma-howard-dermatology.writeupp.com/'
#driverPath = 'C:/Users/Justin Baron/Desktop/Clinics/geckodriver.exe'
driverPath = 'C:/Users/Aliwid/OneDrive/Desktop/Clinics/geckodriver.exe'
thirdURL = writeUppURL + '/admin/thirdparties.aspx'
loginURL = 'https://portal.writeupp.com/login'
patientsURL = writeUppURL + '/admin/data-management/patients.aspx'
activityURL = writeUppURL + '/contactsbydate.aspx'
patientsByInsurer = writeUppURL + '/patientsbythirdparty.aspx'
userName = 'aliwid5@gmail.com'
password = 'Melanoma1!'
testWUID = 'WU1191771'
#wd = 'C:\\Users\\Justin Baron\\Desktop\\Clinics'
wd = 'C:\\Users\\Aliwid\\OneDrive\\Desktop\\Clinics'

HospitalSheetDirectory = wd+'\\New hospital sheets'
template_file = 'Tetbury Blank.xlsx'
downloadDirectory = wd
dr_list = ['Doctor','Dr','Dr.','Doctor,','Dr,','Dr.,']
wu_activity_filename = 'Activity by date.csv'
activity_filename = 'Activity.csv'



profile = webdriver.FirefoxProfile()
profile.set_preference('browser.download.folderList', 2)
profile.set_preference('browser.download.manager.showWhenStarting',False)
profile.set_preference('browser.download.dir', downloadDirectory)
profile.set_preference('browser.helperApps.neverAsk.saveToDisk','text/csv')
driver = webdriver.Firefox(executable_path = driverPath,firefox_profile=profile)

driver.implicitly_wait(10)

def goforit():
    loginWriteupp()
    setup_folder()
    getActivity()
    getInsuranceCompanies()
    process_patients()
    finishUp()
    alldone()


# GUI It all starts here
# Get the login password and billing file suffix
root = Tk()

label_1 = Label(root, text = 'Password:')
label_4 = Label(root, text = 'Clinic Folder' )
label_5 = Label(root, text = version_no)

entryPassword = Entry(root)
entryFolder = Entry(root)

label_1.grid(row=0)
label_4.grid(row=3)
label_5.grid(row=5)

entryPassword.grid(row=0,column = 1)
entryPassword.insert(0,password) # Display the password in the code
entryFolder.grid(row=3,column = 1)

submitButton1 = Button(root, text = 'Press to submit',command=goforit) #live
#submitButton1 = Button(root, text = 'Press to submit',command=testrun) #test

submitButton1.grid(row=4,column=1)

root.mainloop()